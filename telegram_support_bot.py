import json
import logging
import os
import sys
import threading
import urllib.request
import uuid
from datetime import datetime, timedelta

# Fix UTF-8 encoding for Windows console
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl نصب نیست. برای پشتیبانی اکسل: pip install openpyxl")

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update, KeyboardButton, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.constants import ChatType
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

# ====== تنظیمات اولیه ======
TOKEN = os.getenv("BOT_TOKEN")
STATE_FILE = "support_state.json"
INTERNS_FILE = "interns.xlsx"
MEDIA_DIR = "media"
DEFAULT_GROUP_CHAT_ID = os.environ.get("TELEGRAM_GROUP_ID")

# ====== شناسه ادمین‌ها (Telegram user ID) ======
# برای پیدا کردن ID خودت به @userinfobot پیام بده
ADMIN_IDS = [
    123456789,   # <-- ID خودت رو اینجا بنویس
    # 987654321, # ادمین دوم (اختیاری)
]

os.makedirs(MEDIA_DIR, exist_ok=True)

# ====== لیست کامل دوره‌ها ======
COURSES = [
    "نقشه خوانی",
    "نتظیم موتور",
    "پارامتر خوانی",
    "مالتی پلکس ایران خودرو",
    "مالتی پلکس فرانسه",
    "مالتی پلکس سایپا",
    "کولر و تهویه مطبوع",
    "استارت و دینام",
    "هیوندا و کیا",
    "ایسیو ۱",
    "ایسیو ۲",
    "تعمیرات نود مالتی پلکس",
    "جک و لیفان",
    "ریمپ با TNM",
    "کاربری TNM",
    "ایمو بلایزر و تعریف ریموت",
    "وینولز",
]

MESSAGES = {
    "welcome": (
        "🔧 به ربات پشتیبانی آموزشی آکادمی SKP خوش آمدید\n\n"
        "این سامانه با هدف ارائه پشتیبانی تخصصی و پاسخگویی به سوالات کارآموزان دوره‌های آموزشی مکانیک و برق خودرو راه‌اندازی شده است.\n\n"
        "📚 نحوه استفاده از ربات:\n\n"
        "1️⃣ ابتدا شماره تلفن خود را ثبت نمایید.\n\n"
        "2️⃣ سپس دوره آموزشی مورد نظر خود را از لیست دوره‌ها انتخاب کنید.\n\n"
        "3️⃣ سوال فنی، آموزشی یا تخصصی خود را به صورت کامل و دقیق ارسال نمایید.\n"
        "   (می‌توانید متن، عکس، ویدیو یا ویس ارسال کنید)\n\n"
        "4️⃣ سوال شما برای استاد مربوطه ارسال خواهد شد.\n\n"
        "5️⃣ پس از بررسی و پاسخگویی، جواب استاد از طریق همین ربات برای شما ارسال می‌شود.\n\n"
        "💡 برای دریافت پاسخ دقیق‌تر، لطفاً هنگام ثبت سوال مواردی مانند مدل خودرو، سال ساخت، نوع سیستم، کد خطا (در صورت وجود) و توضیحات کامل مشکل را ذکر نمایید.\n\n"
        "⏱ زمان پاسخگویی بسته به نوع سوال، حجم درخواست‌ها و زمان حضور اساتید ممکن است از چند دقیقه تا حداکثر ۲۴ ساعت کاری متغیر باشد.\n\n"
        "☎️ تلفن شرکت:\n021-63002000\n\n"
        "🌐 وب‌سایت:\nhttps://skppart.com/\n\n"
        "از اعتماد شما به آکادمی SKP سپاسگزاریم.\n\n"
    ),
    "group_set": "گروه پشتیبانی ثبت شد.\nID گروه: {group_id}",
    "setgroup_private": "این دستور را داخل گروه پشتیبانی اجرا کنید.",
    "course_selected": "دوره انتخاب شده: {course}\n\nحالا سوال خود را ارسال کن (متن، عکس، ویدیو، ویس، فایل و غیره).",
    "need_course": "ابتدا باید دوره را انتخاب کنید. /start را ارسال کنید.",
    "group_not_set": "گروه پشتیبانی تنظیم نشده است. ابتدا /setgroup را در گروه پشتیبانی اجرا کنید.",
    "question_sent": "سوال شما ثبت شد و به گروه اساتید ارسال شد. به زودی پاسخ دریافت می‌کنید.",
    "question_send_failed": "ارسال سوال به گروه موفق نبود.",
    "selected_answer": "✅ این سوال توسط {teacher} انتخاب شد. لطفاً جواب را در چت خصوصی با ربات ارسال کنید.",
    "teacher_private_question": (
        "شما سوال را برای پاسخ انتخاب کردید:\n"
        "👨‍🎓 دانشجو: {student_name}\n"
        "📚 دوره: {course}\n"
        "سوال: {question}\n\n"
        "لطفاً پاسخ خود را هم‌اکنون در این چت ارسال کنید."
    ),
    "teacher_private_error": "خطا: ربات نمی‌تواند به استاد پیام خصوصی ارسال کند.",
    "not_related_post": "❌ این سوال مربوط به دوره انتخاب‌شده نیست.",
    "student_not_related": "سلام! یکی از اساتید اعلام کرده این سوال مربوط به دوره انتخاب‌شده نیست.",
    "no_pending_question": "شما سوالی برای پاسخ دادن ندارید. ابتدا در گروه سوال را انتخاب کنید.",
    "question_not_found": "سوال پیدا نشد یا قبلاً بسته شده است.",
    "answer_sent": "پاسخ شما با موفقیت به دانشجو ارسال شد.",
    "answer_send_failed": "خطا در ارسال پاسخ به دانشجو.",
    "unknown": "لطفاً از دستور /start برای شروع استفاده کنید.",
    "question_answered_group": "✅ این سوال توسط {teacher} پاسخ داده شد.",
    "not_authorized": (
        "⛔️ شما مجاز به استفاده از این سامانه نیستید.\n\n"
        "شماره‌ی ثبت شده شما در سیستم یافت نشد.\n"
        "برای اطلاعات بیشتر با آکادمی SKP تماس بگیرید:\n"
        "☎️ 021-63002000"
    ),
    "not_admin": "⛔️ شما دسترسی ادمین ندارید.",
    "admin_help": (
        "🛠 دستورات ادمین:\n\n"
        "/adduser 09xxxxxxxxx دوره۱|دوره۲\n"
        "   ➕ افزودن یا ویرایش کاراموز\n\n"
        "/removeuser 09xxxxxxxxx\n"
        "   ➖ حذف کاراموز\n\n"
        "/listusers\n"
        "   📋 نمایش همه کاراموزان\n\n"
        "/getexcel\n"
        "   📥 دریافت فایل اکسل\n\n"
        "📤 آپلود اکسل:\n"
        "   فایل interns.xlsx را برای ربات بفرستید"
    ),
}

STUDENT_PHONE, STUDENT_COURSE, STUDENT_QUESTION = range(3)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.DEBUG,
)
logger = logging.getLogger(__name__)

status = {
    "start_time": None,
    "connected": False,
    "disconnect_count": 0,
    "last_check": None,
    "last_error": None,
}

# =====================================================
# ====== توابع مدیریت دیتابیس اکسل کاراموزان ======
# =====================================================

def normalize_phone(phone: str) -> str:
    """نرمال‌سازی شماره تلفن به فرمت 09XXXXXXXXX"""
    phone = str(phone).strip().replace(" ", "").replace("-", "").replace("+", "")
    if phone.startswith("98") and len(phone) == 12:
        phone = "0" + phone[2:]
    if phone.startswith("9") and len(phone) == 10:
        phone = "0" + phone
    return phone


def load_interns_db() -> dict:
    """
    بارگذاری دیتابیس کاراموزان از فایل اکسل.
    فرمت: ستون A = شماره موبایل، ستون B = دوره‌ها با | جدا شده
    مثال B: نقشه خوانی|ایسیو ۱|مالتی پلکس ایران خودرو
    """
    if not EXCEL_AVAILABLE:
        logger.warning("openpyxl نصب نیست")
        return {}
    if not os.path.exists(INTERNS_FILE):
        return {}
    try:
        wb = openpyxl.load_workbook(INTERNS_FILE)
        ws = wb.active
        db = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            phone_raw = row[0]
            courses_raw = row[1]
            if not phone_raw:
                continue
            phone = normalize_phone(str(phone_raw))
            courses = [c.strip() for c in str(courses_raw).split("|") if c.strip()] if courses_raw else []
            db[phone] = courses
        logger.info("دیتابیس از اکسل: %d کاراموز", len(db))
        return db
    except Exception as e:
        logger.error("خطا در خواندن اکسل: %s", e)
        return {}


def save_interns_db(db: dict) -> bool:
    """ذخیره دیتابیس در فایل اکسل"""
    if not EXCEL_AVAILABLE:
        return False
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "کاراموزان"
        ws["A1"] = "شماره موبایل"
        ws["B1"] = "دوره‌ها (با | جدا کنید)"
        ws["A1"].font = openpyxl.styles.Font(bold=True)
        ws["B1"].font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 70
        for i, (phone, courses) in enumerate(db.items(), start=2):
            ws.cell(row=i, column=1, value=phone)
            ws.cell(row=i, column=2, value="|".join(courses))
        wb.save(INTERNS_FILE)
        return True
    except Exception as e:
        logger.error("خطا در ذخیره اکسل: %s", e)
        return False


def get_allowed_courses(phone: str):
    """برمیگردونه لیست دوره‌های مجاز یا None اگر نبود"""
    db = load_interns_db()
    return db.get(normalize_phone(phone))


def add_intern(phone: str, courses: list) -> bool:
    db = load_interns_db()
    db[normalize_phone(phone)] = courses
    return save_interns_db(db)


def remove_intern(phone: str) -> bool:
    db = load_interns_db()
    normalized = normalize_phone(phone)
    if normalized not in db:
        return False
    del db[normalized]
    return save_interns_db(db)


def create_sample_excel_if_missing():
    """اگر فایل اکسل وجود نداشت، نمونه بساز"""
    if not EXCEL_AVAILABLE or os.path.exists(INTERNS_FILE):
        return
    sample = {
        "09121234567": ["نقشه خوانی", "ایسیو ۱"],
        "09351234567": ["مالتی پلکس ایران خودرو", "هیوندا و کیا"],
    }
    save_interns_db(sample)
    logger.info("فایل اکسل نمونه ساخته شد: %s", INTERNS_FILE)

# =====================================================
# ====== توابع کمکی ======
# =====================================================

def format_timedelta(delta: timedelta) -> str:
    total = int(delta.total_seconds())
    h, r = divmod(total, 3600)
    m, s = divmod(r, 60)
    if h:
        return f"{h}h {m:02d}m {s:02d}s"
    if m:
        return f"{m}m {s:02d}s"
    return f"{s}s"


def clear_console() -> None:
    if os.name == "nt":
        os.system("cls")
    else:
        sys.stdout.write("\033[2J\033[H")
        sys.stdout.flush()


def get_connection_status():
    url = f"https://api.telegram.org/bot{TOKEN}/getMe"
    try:
        with urllib.request.urlopen(url, timeout=10) as r:
            data = json.load(r)
        return bool(data.get("ok")), None
    except Exception as e:
        return False, str(e)


def display_status(final: bool = False) -> None:
    clear_console()
    now = datetime.now()
    uptime = format_timedelta(now - (status["start_time"] or now))
    print("============================================")
    print("    Telegram Support Bot - وضعیت اتصال")
    print("============================================")
    print(f"وضعیت: {'✅ متصل' if status['connected'] else '❌ قطع'}")
    print(f"زمان اجرا: {uptime}")
    print(f"تعداد قطعی‌ها: {status['disconnect_count']}")
    print(f"آخرین بررسی: {status['last_check'].strftime('%Y-%m-%d %H:%M:%S') if status['last_check'] else '-'}")
    if status["last_error"]:
        print(f"خطای آخر: {status['last_error']}")
    print("--------------------------------------------")
    print("برای توقف: Ctrl+C")
    if final:
        print("ربات متوقف شده است.")
    sys.stdout.flush()


def status_monitor(stop_event: threading.Event) -> None:
    while not stop_event.is_set():
        connected, error = get_connection_status()
        now = datetime.now()
        if status["start_time"] is None:
            status["start_time"] = now
        if status["connected"] and not connected:
            status["disconnect_count"] += 1
        status["connected"] = connected
        status["last_check"] = now
        status["last_error"] = error
        display_status()
        if stop_event.wait(5):
            break
    display_status(final=True)


_state_default = {
    "group_chat_id": DEFAULT_GROUP_CHAT_ID,
    "questions": {},
    "teacher_pending": {},
}


def load_state() -> dict:
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.warning("بارگذاری state موفق نبود: %s", e)
    return dict(_state_default)


def save_state(data: dict) -> None:
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.error("ذخیره state خطا: %s", e)


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def build_course_keyboard(allowed_courses=None) -> list:
    courses = allowed_courses if allowed_courses is not None else COURSES
    return [[InlineKeyboardButton(c, callback_data=f"course:{c}")] for c in courses]

# =====================================================
# ====== دستورات ادمین ======
# =====================================================

async def admin_help(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_admin(update.effective_user.id):
        await update.message.reply_text(MESSAGES["not_admin"])
        return
    await update.message.reply_text(MESSAGES["admin_help"])


async def admin_add_user(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    /adduser 09121234567 نقشه خوانی|ایسیو ۱|مالتی پلکس ایران خودرو
    """
    if not is_admin(update.effective_user.id):
        await update.message.reply_text(MESSAGES["not_admin"])
        return
    args = context.args
    if len(args) < 2:
        await update.message.reply_text(
            "فرمت اشتباه!\n\nنمونه:\n/adduser 09121234567 نقشه خوانی|ایسیو ۱\n\n"
            "دوره‌ها را با | از هم جدا کنید."
        )
        return
    phone = normalize_phone(args[0])
    courses_str = " ".join(args[1:])
    courses = [c.strip() for c in courses_str.split("|") if c.strip()]
    invalid = [c for c in courses if c not in COURSES]
    if invalid:
        courses_list = "\n".join(f"• {c}" for c in COURSES)
        await update.message.reply_text(
            f"⚠️ دوره‌های نامعتبر:\n{', '.join(invalid)}\n\nدوره‌های مجاز:\n{courses_list}"
        )
        return
    if add_intern(phone, courses):
        await update.message.reply_text(
            f"✅ ثبت/ویرایش شد:\n📞 {phone}\n📚 دوره‌ها:\n" +
            "\n".join(f"  ✅ {c}" for c in courses)
        )
    else:
        await update.message.reply_text("❌ خطا. مطمئن شوید openpyxl نصب است:\npip install openpyxl")


async def admin_remove_user(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """/removeuser 09121234567"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text(MESSAGES["not_admin"])
        return
    args = context.args
    if not args:
        await update.message.reply_text("فرمت: /removeuser 09121234567")
        return
    phone = normalize_phone(args[0])
    if remove_intern(phone):
        await update.message.reply_text(f"✅ کاراموز {phone} حذف شد.")
    else:
        await update.message.reply_text(f"⚠️ شماره {phone} در سیستم یافت نشد.")


async def admin_list_users(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """/listusers"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text(MESSAGES["not_admin"])
        return
    db = load_interns_db()
    if not db:
        await update.message.reply_text("📋 هیچ کاراموزی ثبت نشده است.")
        return
    lines = [f"📋 لیست کاراموزان ({len(db)} نفر):\n"]
    for phone, courses in db.items():
        lines.append(f"📞 {phone}\n   {' | '.join(courses) if courses else 'بدون دوره'}\n")
    text = "\n".join(lines)
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await update.message.reply_text(chunk)


async def admin_get_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """/getexcel - ارسال فایل اکسل"""
    if not is_admin(update.effective_user.id):
        await update.message.reply_text(MESSAGES["not_admin"])
        return
    if not os.path.exists(INTERNS_FILE):
        await update.message.reply_text("❌ فایل اکسل وجود ندارد.")
        return
    try:
        with open(INTERNS_FILE, "rb") as f:
            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=f,
                filename="interns.xlsx",
                caption="📊 فایل اکسل کاراموزان"
            )
    except Exception as e:
        await update.message.reply_text(f"❌ خطا: {e}")


async def admin_upload_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """آپلود فایل اکسل توسط ادمین"""
    if not is_admin(update.effective_user.id):
        return
    doc = update.message.document
    if not doc or not (doc.file_name or "").endswith(".xlsx"):
        await update.message.reply_text(
            "⚠️ فقط فایل .xlsx قبول می‌شود.\n"
            "فرمت: ستون A = شماره موبایل، ستون B = دوره‌ها با | جدا شده"
        )
        return
    try:
        file = await context.bot.get_file(doc.file_id)
        await file.download_to_drive(INTERNS_FILE)
        db = load_interns_db()
        await update.message.reply_text(
            f"✅ فایل اکسل بارگذاری شد!\n📊 تعداد کاراموزان: {len(db)} نفر\n\nبرای مشاهده: /listusers"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ خطا: {e}")

# =====================================================
# ====== هندلرهای اصلی ربات ======
# =====================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.effective_chat.type != ChatType.PRIVATE:
        await update.message.reply_text("لطفاً در چت خصوصی با من /start را ارسال کنید.")
        return ConversationHandler.END
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("شروع ثبت درخواست", callback_data="start_register")]])
    await update.message.reply_text(MESSAGES["welcome"], reply_markup=keyboard)
    return STUDENT_PHONE


async def set_group(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_chat.type in (ChatType.GROUP, ChatType.SUPERGROUP):
        state_data = load_state()
        state_data["group_chat_id"] = update.effective_chat.id
        save_state(state_data)
        await update.message.reply_text(MESSAGES["group_set"].format(group_id=update.effective_chat.id))
    else:
        await update.message.reply_text(MESSAGES["setgroup_private"])


async def course_selected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    course = query.data.split(":", 1)[1]
    context.user_data["selected_course"] = course
    await query.message.edit_text(MESSAGES["course_selected"].format(course=course))
    return STUDENT_QUESTION


async def ask_again_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    context.user_data.pop("selected_course", None)
    allowed_courses = context.user_data.get("allowed_courses")
    try:
        await context.bot.send_message(
            chat_id=query.from_user.id,
            text="دوره مورد نظر را انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup(build_course_keyboard(allowed_courses)),
        )
    except Exception:
        try:
            await query.message.edit_text(
                "دوره مورد نظر را انتخاب کنید:",
                reply_markup=InlineKeyboardMarkup(build_course_keyboard(allowed_courses)),
            )
        except Exception:
            pass
    return STUDENT_COURSE


async def start_register_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    kb = ReplyKeyboardMarkup([[KeyboardButton("ارسال شماره", request_contact=True)]], one_time_keyboard=True, resize_keyboard=True)
    try:
        await context.bot.send_message(
            chat_id=query.from_user.id,
            text="لطفاً شماره تلفن خود را ارسال کنید (از طریق دکمه یا تایپ شماره).",
            reply_markup=kb
        )
    except Exception:
        pass
    return STUDENT_PHONE


async def receive_contact(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    phone = None
    if update.message.contact and update.message.contact.phone_number:
        phone = update.message.contact.phone_number
    elif update.message.text:
        txt = update.message.text.strip()
        if any(ch.isdigit() for ch in txt):
            phone = txt
    if not phone:
        await update.message.reply_text("شماره تلفن نامعتبر است. لطفاً مجدداً امتحان کنید.")
        return STUDENT_PHONE

    allowed_courses = get_allowed_courses(phone)
    if allowed_courses is None:
        logger.info("شماره %s مجاز نیست", phone)
        await update.message.reply_text(MESSAGES["not_authorized"], reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END

    normalized = normalize_phone(phone)
    state_data = load_state()
    users = state_data.get("users") or {}
    users[str(user.id)] = {"phone": normalized, "allowed_courses": allowed_courses, "name": user.full_name}
    state_data["users"] = users
    save_state(state_data)
    context.user_data["phone"] = normalized
    context.user_data["allowed_courses"] = allowed_courses

    try:
        await context.bot.send_message(
            chat_id=user.id,
            text=f"✅ شماره شما تأیید شد: {normalized}\n\n📚 دوره‌های در دسترس شما:",
            reply_markup=ReplyKeyboardRemove()
        )
        await context.bot.send_message(
            chat_id=user.id,
            text="لطفاً دوره مورد نظر را انتخاب کنید:",
            reply_markup=InlineKeyboardMarkup(build_course_keyboard(allowed_courses))
        )
    except Exception:
        pass
    return STUDENT_COURSE


async def receive_question(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    course = context.user_data.get("selected_course")
    question_text = media_file_id = media_type = None

    if update.message.text:
        question_text = update.message.text.strip()
    elif update.message.photo:
        question_text = update.message.caption or "عکس ارسال شده"
        media_file_id = update.message.photo[-1].file_id
        media_type = "photo"
    elif update.message.video:
        question_text = update.message.caption or "ویدیو ارسال شده"
        media_file_id = update.message.video.file_id
        media_type = "video"
    elif update.message.voice:
        question_text = update.message.caption or "ویس ارسال شده"
        media_file_id = update.message.voice.file_id
        media_type = "voice"
    elif update.message.document:
        question_text = update.message.caption or "📎 فایل ارسال شده"
        media_file_id = update.message.document.file_id
        media_type = "document"
    elif update.message.audio:
        question_text = update.message.caption or "🎵 آهنگ ارسال شده"
        media_file_id = update.message.audio.file_id
        media_type = "audio"
    elif update.message.animation:
        question_text = update.message.caption or "🎬 GIF ارسال شده"
        media_file_id = update.message.animation.file_id
        media_type = "animation"
    else:
        await update.message.reply_text("لطفاً متن، عکس، ویدیو، ویس یا فایل ارسال کنید.")
        return STUDENT_QUESTION

    if not course:
        await update.message.reply_text(MESSAGES["need_course"])
        return ConversationHandler.END

    state_data = load_state()
    group_chat_id = state_data.get("group_chat_id") or DEFAULT_GROUP_CHAT_ID
    if not group_chat_id:
        await update.message.reply_text(MESSAGES["group_not_set"])
        return ConversationHandler.END

    question_id = str(uuid.uuid4())
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    state_data["questions"][question_id] = {
        "student_id": user.id,
        "student_name": user.full_name,
        "course": course,
        "question": question_text,
        "status": "open",
        "created_at": now,
        "group_message_id": None,
        "assigned_teacher_id": None,
        "assigned_teacher_name": None,
        "answer": None,
        "media_file_id": media_file_id,
        "media_type": media_type,
    }

    phone = (state_data.get("users") or {}).get(str(user.id), {}).get("phone") or context.user_data.get("phone")
    group_message = (
        f"سوال جدید ثبت شد:\n"
        f"👨‍🎓 دانشجو: {user.full_name}\n"
        + (f"📞 {phone}\n" if phone else "")
        + f"📚 دوره: {course}\n"
        f"🕒 زمان: {now}\n\nسوال: {question_text}"
    )
    keyboard = [
        [InlineKeyboardButton("✅ پاسخ می‌دهم", callback_data=f"answer:{question_id}")],
        [InlineKeyboardButton("❌ مربوط به این دوره نیست", callback_data=f"not_related:{question_id}")],
    ]
    try:
        msg = await context.bot.send_message(
            chat_id=group_chat_id, text=group_message,
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        state_data["questions"][question_id]["group_message_id"] = msg.message_id
        if media_file_id and media_type:
            try:
                if media_type == "photo":
                    await context.bot.send_photo(chat_id=group_chat_id, photo=media_file_id, caption="📷 عکس سوال")
                elif media_type == "video":
                    await context.bot.send_video(chat_id=group_chat_id, video=media_file_id, caption="🎥 ویدیو سوال")
                elif media_type == "voice":
                    await context.bot.send_voice(chat_id=group_chat_id, voice=media_file_id, caption="🎙️ ویس سوال")
                elif media_type == "document":
                    await context.bot.send_document(chat_id=group_chat_id, document=media_file_id, caption="📎 فایل سوال")
                elif media_type == "audio":
                    await context.bot.send_audio(chat_id=group_chat_id, audio=media_file_id, caption="🎵 آهنگ سوال")
                elif media_type == "animation":
                    await context.bot.send_animation(chat_id=group_chat_id, animation=media_file_id, caption="🎬 GIF سوال")
            except Exception as e:
                logger.error("خطا رسانه به گروه: %s", e)
        save_state(state_data)
        await update.message.reply_text(MESSAGES["question_sent"])
        try:
            await context.bot.send_message(
                chat_id=user.id, text="در صورت داشتن سوال مجدد:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("ارسال سوال جدید", callback_data="ask_again")]]),
            )
        except Exception:
            pass
    except Exception as e:
        logger.error("خطا ارسال سوال: %s", e)
        await update.message.reply_text(MESSAGES["question_send_failed"])
    return ConversationHandler.END


async def group_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    action, question_id = query.data.split(":", 1)
    state_data = load_state()
    question = state_data["questions"].get(question_id)
    if not question:
        await query.edit_message_text("این سوال موجود نیست.")
        return
    if question["status"] != "open":
        await query.answer("این سوال قبلاً بررسی شده است.", show_alert=True)
        return
    if action == "answer":
        teacher = update.effective_user
        question["status"] = "assigned"
        question["assigned_teacher_id"] = teacher.id
        question["assigned_teacher_name"] = teacher.full_name
        state_data["teacher_pending"][str(teacher.id)] = question_id
        save_state(state_data)
        await query.edit_message_text(
            query.message.text + "\n\n" + MESSAGES["selected_answer"].format(teacher=teacher.full_name)
        )
        try:
            await context.bot.send_message(
                chat_id=teacher.id,
                text=MESSAGES["teacher_private_question"].format(
                    student_name=question["student_name"],
                    course=question["course"],
                    question=question["question"],
                ),
            )
            if question.get("media_file_id") and question.get("media_type"):
                mt, fid = question["media_type"], question["media_file_id"]
                if mt == "photo":
                    await context.bot.send_photo(chat_id=teacher.id, photo=fid, caption="📷 عکس سوال")
                elif mt == "video":
                    await context.bot.send_video(chat_id=teacher.id, video=fid, caption="🎥 ویدیو سوال")
                elif mt == "voice":
                    await context.bot.send_voice(chat_id=teacher.id, voice=fid, caption="🎙️ ویس سوال")
                elif mt == "document":
                    await context.bot.send_document(chat_id=teacher.id, document=fid, caption="📎 فایل سوال")
                elif mt == "audio":
                    await context.bot.send_audio(chat_id=teacher.id, audio=fid, caption="🎵 آهنگ سوال")
                elif mt == "animation":
                    await context.bot.send_animation(chat_id=teacher.id, animation=fid, caption="🎬 GIF سوال")
        except Exception as e:
            logger.error("خطا ارسال به استاد: %s", e)
            await query.message.reply_text(MESSAGES["teacher_private_error"])
    elif action == "not_related":
        question["status"] = "not_related"
        save_state(state_data)
        await query.edit_message_text(query.message.text + "\n\n" + MESSAGES["not_related_post"])
        try:
            await context.bot.send_message(chat_id=question["student_id"], text=MESSAGES["student_not_related"])
        except Exception as e:
            logger.error("خطا not_related: %s", e)


async def teacher_reply(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    # اگر ادمین فایل xlsx فرستاد، پردازش کن
    if update.message.document and is_admin(update.effective_user.id):
        if (update.message.document.file_name or "").endswith(".xlsx"):
            await admin_upload_excel(update, context)
            return

    teacher_id = str(update.effective_user.id)
    state_data = load_state()
    pending = state_data["teacher_pending"].get(teacher_id)
    if not pending:
        await update.message.reply_text(MESSAGES["no_pending_question"])
        return

    answer_text = answer_media_file_id = answer_media_type = None
    if update.message.text:
        answer_text = update.message.text.strip()
    elif update.message.photo:
        answer_text = update.message.caption or "📷 عکس پاسخ"
        answer_media_file_id = update.message.photo[-1].file_id
        answer_media_type = "photo"
    elif update.message.video:
        answer_text = update.message.caption or "🎥 ویدیو پاسخ"
        answer_media_file_id = update.message.video.file_id
        answer_media_type = "video"
    elif update.message.voice:
        answer_text = update.message.caption or "🎙️ ویس پاسخ"
        answer_media_file_id = update.message.voice.file_id
        answer_media_type = "voice"
    elif update.message.audio:
        answer_text = update.message.caption or "🎵 آهنگ پاسخ"
        answer_media_file_id = update.message.audio.file_id
        answer_media_type = "audio"
    elif update.message.document:
        answer_text = update.message.caption or "📎 فایل پاسخ"
        answer_media_file_id = update.message.document.file_id
        answer_media_type = "document"
    elif update.message.animation:
        answer_text = update.message.caption or "🎬 GIF پاسخ"
        answer_media_file_id = update.message.animation.file_id
        answer_media_type = "animation"
    else:
        await update.message.reply_text("لطفاً متن یا رسانه‌ای برای پاسخ ارسال کنید.")
        return

    question = state_data["questions"].get(pending)
    if not question:
        await update.message.reply_text(MESSAGES["question_not_found"])
        return

    question["status"] = "answered"
    question["answer"] = answer_text
    question["answer_media_file_id"] = answer_media_file_id
    question["answer_media_type"] = answer_media_type
    state_data["teacher_pending"].pop(teacher_id, None)
    save_state(state_data)

    student_msg = (
        f"✅ پاسخ استاد:\n📚 دوره: {question['course']}\n"
        f"👨‍🏫 استاد: {question['assigned_teacher_name']}\n\nپاسخ: {answer_text}"
    )
    try:
        await context.bot.send_message(
            chat_id=question["student_id"], text=student_msg,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("سوالی ندارم", callback_data=f"no_more:{pending}")],
                [InlineKeyboardButton("باز سوال دارم", callback_data=f"ask_more:{pending}")],
            ])
        )
        if answer_media_file_id and answer_media_type:
            mt, fid, sid = answer_media_type, answer_media_file_id, question["student_id"]
            if mt == "photo":
                await context.bot.send_photo(chat_id=sid, photo=fid, caption="📷 عکس پاسخ")
            elif mt == "video":
                await context.bot.send_video(chat_id=sid, video=fid, caption="🎥 ویدیو پاسخ")
            elif mt == "voice":
                await context.bot.send_voice(chat_id=sid, voice=fid, caption="🎙️ ویس پاسخ")
            elif mt == "audio":
                await context.bot.send_audio(chat_id=sid, audio=fid, caption="🎵 آهنگ پاسخ")
            elif mt == "document":
                await context.bot.send_document(chat_id=sid, document=fid, caption="📎 فایل پاسخ")
            elif mt == "animation":
                await context.bot.send_animation(chat_id=sid, animation=fid, caption="🎬 GIF پاسخ")
        await update.message.reply_text(MESSAGES["answer_sent"])
    except Exception as e:
        logger.error("خطا ارسال پاسخ: %s", e)
        await update.message.reply_text(MESSAGES["answer_send_failed"])

    group_chat_id = state_data.get("group_chat_id") or DEFAULT_GROUP_CHAT_ID
    if group_chat_id and question.get("group_message_id"):
        try:
            await context.bot.edit_message_text(
                chat_id=group_chat_id,
                message_id=question["group_message_id"],
                text=f"{question['question']}\n\n" + MESSAGES["question_answered_group"].format(teacher=question["assigned_teacher_name"]),
            )
        except Exception as e:
            logger.warning("خطا update گروه: %s", e)


async def post_answer_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    action, question_id = query.data.split(":", 1)
    state_data = load_state()
    question = state_data["questions"].get(question_id)
    if not question:
        try:
            await query.edit_message_text("خطا: سوال یافت نشد.")
        except Exception:
            pass
        return ConversationHandler.END
    if action == "no_more":
        try:
            await context.bot.send_poll(
                chat_id=question["student_id"],
                question="نظرسنجی: کیفیت پاسخ پشتیبانی؟",
                options=["خیلی خوب", "خوب", "متوسط", "ضعیف"],
                is_anonymous=False,
            )
            await context.bot.send_message(
                chat_id=question["student_id"], text="اگر می‌خواهید دوباره استفاده کنید:",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("استارت مجدد", callback_data="restart_bot")]]),
            )
        except Exception as e:
            logger.error("خطا نظرسنجی: %s", e)
        return ConversationHandler.END
    elif action == "ask_more":
        allowed_courses = (state_data.get("users") or {}).get(str(question["student_id"]), {}).get("allowed_courses")
        try:
            await context.bot.send_message(
                chat_id=question["student_id"], text="دوره مورد نظر رو انتخاب کنید:",
                reply_markup=InlineKeyboardMarkup(build_course_keyboard(allowed_courses))
            )
        except Exception:
            pass
        return STUDENT_COURSE


async def restart_bot_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    try:
        await query.message.reply_text(
            MESSAGES["welcome"],
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("شروع ثبت درخواست", callback_data="start_register")]]),
        )
    except Exception:
        pass


async def unknown(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(MESSAGES["unknown"])

# =====================================================
# ====== main ======
# =====================================================

def main() -> None:
    create_sample_excel_if_missing()
    app = ApplicationBuilder().token(TOKEN).build()

    # دستورات ادمین (باید قبل از conv_handler باشند)
    app.add_handler(CommandHandler("admin", admin_help))
    app.add_handler(CommandHandler("adduser", admin_add_user))
    app.add_handler(CommandHandler("removeuser", admin_remove_user))
    app.add_handler(CommandHandler("listusers", admin_list_users))
    app.add_handler(CommandHandler("getexcel", admin_get_excel))

    conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            CallbackQueryHandler(post_answer_callback, pattern=r"^ask_more:"),
        ],
        states={
            STUDENT_PHONE: [
                CallbackQueryHandler(start_register_callback, pattern=r"^start_register$"),
                MessageHandler(filters.CONTACT | (filters.TEXT & ~filters.COMMAND), receive_contact),
            ],
            STUDENT_COURSE: [CallbackQueryHandler(course_selected, pattern=r"^course:")],
            STUDENT_QUESTION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_question),
                MessageHandler(filters.PHOTO, receive_question),
                MessageHandler(filters.VIDEO, receive_question),
                MessageHandler(filters.VOICE, receive_question),
                MessageHandler(filters.Document.ALL, receive_question),
                MessageHandler(filters.AUDIO, receive_question),
                MessageHandler(filters.ANIMATION, receive_question),
            ],
        },
        fallbacks=[CommandHandler("start", start)],
    )

    app.add_handler(conv_handler)
    app.add_handler(CallbackQueryHandler(post_answer_callback, pattern=r"^(no_more|ask_more):"))
    app.add_handler(CallbackQueryHandler(restart_bot_callback, pattern=r"^restart_bot$"))
    app.add_handler(CommandHandler("setgroup", set_group))
    app.add_handler(CallbackQueryHandler(group_callback, pattern=r"^(answer|not_related):"))
    app.add_handler(CallbackQueryHandler(course_selected, pattern=r"^course:"))
    app.add_handler(CallbackQueryHandler(ask_again_callback, pattern=r"^ask_again$"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, teacher_reply))
    app.add_handler(MessageHandler(filters.PHOTO, teacher_reply))
    app.add_handler(MessageHandler(filters.VIDEO, teacher_reply))
    app.add_handler(MessageHandler(filters.VOICE, teacher_reply))
    app.add_handler(MessageHandler(filters.AUDIO, teacher_reply))
    app.add_handler(MessageHandler(filters.Document.ALL, teacher_reply))
    app.add_handler(MessageHandler(filters.ANIMATION, teacher_reply))
    app.add_handler(MessageHandler(filters.ALL, unknown))

    stop_event = threading.Event()
    monitor_thread = threading.Thread(target=status_monitor, args=(stop_event,), daemon=True)
    monitor_thread.start()

    logger.info("ربات شروع به کار کرد...")
    try:
        app.run_polling()
    finally:
        stop_event.set()
        monitor_thread.join(timeout=5)


if __name__ == "__main__":
    main()
